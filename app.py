from flask import Flask, render_template, request, redirect, send_file
import sqlite3
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
import os
from datetime import date
import calendar


app = Flask(__name__)

DATABASE = "attendance.db"

SITES = ["TSM", "TSK", "NINL"]

VALID_STATUSES = ["P", "A", "W/O", "NH", "FL"]

FY_START_YEAR = 2026
FY_END_YEAR = 2027

EPF_RATE = 12.0
ESIC_RATE = 0.75
OT_MULTIPLIER = 2.0


# =========================================================
# DATABASE
# =========================================================

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def add_column_if_missing(
    conn,
    table_name,
    column_name,
    definition
):

    columns = conn.execute(
        f"PRAGMA table_info({table_name})"
    ).fetchall()

    names = [
        column["name"]
        for column in columns
    ]

    if column_name not in names:

        conn.execute(
            f"""
            ALTER TABLE {table_name}
            ADD COLUMN {column_name} {definition}
            """
        )


def init_db():

    conn = get_db()

    # -----------------------------------------------------
    # EMPLOYEES
    # -----------------------------------------------------

    conn.execute("""
        CREATE TABLE IF NOT EXISTS employees (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            name TEXT NOT NULL,

            employee_code TEXT NOT NULL UNIQUE,

            designation TEXT NOT NULL,

            skill TEXT NOT NULL DEFAULT '',

            site TEXT NOT NULL,

            basic REAL DEFAULT 0,

            gross REAL DEFAULT 0

        )
    """)

    # -----------------------------------------------------
    # ATTENDANCE
    # -----------------------------------------------------

    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            employee_id INTEGER NOT NULL,

            attendance_date TEXT NOT NULL,

            status TEXT NOT NULL,

            ot_hours REAL DEFAULT 0,

            UNIQUE(
                employee_id,
                attendance_date
            )

        )
    """)

    # -----------------------------------------------------
    # SKILL RATES
    # -----------------------------------------------------

    conn.execute("""
        CREATE TABLE IF NOT EXISTS skill_rates (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            skill TEXT NOT NULL UNIQUE,

            rate REAL DEFAULT 0

        )
    """)

    # -----------------------------------------------------
    # SALARY DEDUCTIONS
    # -----------------------------------------------------

    conn.execute("""
        CREATE TABLE IF NOT EXISTS salary_deductions (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            employee_id INTEGER NOT NULL,

            month INTEGER NOT NULL,

            year INTEGER NOT NULL,

            profession_tax REAL DEFAULT 0,

            advance_deduction REAL DEFAULT 0,

            other_deduction REAL DEFAULT 0,

            UNIQUE(
                employee_id,
                month,
                year
            )

        )
    """)

    # -----------------------------------------------------
    # OLD DATABASE COMPATIBILITY
    # -----------------------------------------------------

    add_column_if_missing(
        conn,
        "employees",
        "skill",
        "TEXT NOT NULL DEFAULT ''"
    )

    add_column_if_missing(
        conn,
        "employees",
        "basic",
        "REAL DEFAULT 0"
    )

    add_column_if_missing(
        conn,
        "employees",
        "gross",
        "REAL DEFAULT 0"
    )

    conn.commit()

    conn.close()


# =========================================================
# HOME
# =========================================================

@app.route("/")
def home():

    return render_template(
        "index.html"
    )


# =========================================================
# CENTRAL MANAGE EMPLOYEE
# =========================================================

@app.route("/manage-employees")
def manage_employees():

    search = request.args.get(
        "search",
        ""
    ).strip()

    selected_site = request.args.get(
        "site",
        "ALL"
    ).strip().upper()

    if (
        selected_site != "ALL"
        and selected_site not in SITES
    ):
        selected_site = "ALL"

    conn = get_db()

    if selected_site == "ALL":

        if search:

            pattern = f"%{search}%"

            employees = conn.execute("""
                SELECT *
                FROM employees
                WHERE
                    name LIKE ?
                    OR employee_code LIKE ?
                    OR designation LIKE ?
                    OR skill LIKE ?
                    OR site LIKE ?
                ORDER BY name
            """, (
                pattern,
                pattern,
                pattern,
                pattern,
                pattern
            )).fetchall()

        else:

            employees = conn.execute("""
                SELECT *
                FROM employees
                ORDER BY name
            """).fetchall()

    else:

        if search:

            pattern = f"%{search}%"

            employees = conn.execute("""
                SELECT *
                FROM employees
                WHERE site = ?
                AND (
                    name LIKE ?
                    OR employee_code LIKE ?
                    OR designation LIKE ?
                    OR skill LIKE ?
                )
                ORDER BY name
            """, (
                selected_site,
                pattern,
                pattern,
                pattern,
                pattern
            )).fetchall()

        else:

            employees = conn.execute("""
                SELECT *
                FROM employees
                WHERE site = ?
                ORDER BY name
            """, (
                selected_site,
            )).fetchall()

    total_employees = conn.execute(
        "SELECT COUNT(*) FROM employees"
    ).fetchone()[0]

    conn.close()

    return render_template(
        "manage_employees.html",
        employees=employees,
        sites=SITES,
        selected_site=selected_site,
        search=search,
        total_employees=total_employees
    )


# =========================================================
# DELETE EMPLOYEE
# =========================================================

@app.route(
    "/delete-employee/<int:employee_id>",
    methods=["POST"]
)
def delete_employee(employee_id):

    conn = get_db()

    employee = conn.execute("""
        SELECT id
        FROM employees
        WHERE id = ?
    """, (
        employee_id,
    )).fetchone()

    if not employee:

        conn.close()

        return "Employee not found."

    conn.execute("""
        DELETE FROM attendance
        WHERE employee_id = ?
    """, (
        employee_id,
    ))

    conn.execute("""
        DELETE FROM salary_deductions
        WHERE employee_id = ?
    """, (
        employee_id,
    ))

    conn.execute("""
        DELETE FROM employees
        WHERE id = ?
    """, (
        employee_id,
    ))

    conn.commit()

    conn.close()

    return redirect(
        "/manage-employees"
    )


# =========================================================
# FILL ATTENDANCE
# =========================================================

@app.route("/fill-attendance")
def fill_attendance():

    return render_template(
        "site.html",
        sites=SITES
    )


# =========================================================
# FINANCIAL YEAR MONTHS
# =========================================================

def get_fy_months():

    return [

        {
            "month": 4,
            "year": 2026,
            "name": "April 2026"
        },

        {
            "month": 5,
            "year": 2026,
            "name": "May 2026"
        },

        {
            "month": 6,
            "year": 2026,
            "name": "June 2026"
        },

        {
            "month": 7,
            "year": 2026,
            "name": "July 2026"
        },

        {
            "month": 8,
            "year": 2026,
            "name": "August 2026"
        },

        {
            "month": 9,
            "year": 2026,
            "name": "September 2026"
        },

        {
            "month": 10,
            "year": 2026,
            "name": "October 2026"
        },

        {
            "month": 11,
            "year": 2026,
            "name": "November 2026"
        },

        {
            "month": 12,
            "year": 2026,
            "name": "December 2026"
        },

        {
            "month": 1,
            "year": 2027,
            "name": "January 2027"
        },

        {
            "month": 2,
            "year": 2027,
            "name": "February 2027"
        },

        {
            "month": 3,
            "year": 2027,
            "name": "March 2027"
        }

    ]


# =========================================================
# SELECT MONTH
# =========================================================

@app.route("/select-month/<site>")
def select_month(site):

    site = site.upper()

    if site not in SITES:

        return "Invalid site."

    return render_template(
        "month.html",
        site=site,
        months=get_fy_months()
    )


def valid_financial_year(
    month,
    year
):

    return (

        month in range(4, 13)
        and year == 2026

    ) or (

        month in range(1, 4)
        and year == 2027

    )


# =========================================================
# DAILY ATTENDANCE - DATES
# =========================================================

@app.route(
    "/attendance-dates/<site>/<int:month>/<int:year>"
)
def attendance_dates(
    site,
    month,
    year
):

    site = site.upper()

    if site not in SITES:

        return "Invalid site."

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    today = date.today()

    total_days = calendar.monthrange(
        year,
        month
    )[1]

    conn = get_db()

    employee_count = conn.execute("""
        SELECT COUNT(*)
        FROM employees
        WHERE site = ?
    """, (
        site,
    )).fetchone()[0]

    conn.close()

    days = []

    for d in range(
        1,
        total_days + 1
    ):

        current = date(
            year,
            month,
            d
        )

        days.append({

            "date": d,

            "day": current.strftime(
                "%A"
            ),

            "short_day": current.strftime(
                "%a"
            ),

            "full_date": current.strftime(
                "%d %B %Y"
            ),

            "future": current > today

        })

    return render_template(
        "attendance_dates.html",
        site=site,
        month=month,
        year=year,
        month_name=calendar.month_name[
            month
        ],
        days=days,
        employee_count=employee_count,
        today=today.strftime(
            "%Y-%m-%d"
        )
    )


# =========================================================
# ATTENDANCE SELECT
# =========================================================

@app.route(
    "/attendance-select/<site>/<int:month>/<int:year>/<int:day>"
)
def attendance_select(
    site,
    month,
    year,
    day
):

    site = site.upper()

    if site not in SITES:

        return "Invalid site."

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    total_days = calendar.monthrange(
        year,
        month
    )[1]

    if day < 1 or day > total_days:

        return "Invalid date."

    selected_date = date(
        year,
        month,
        day
    )

    if selected_date > date.today():

        return "Future date attendance is locked."

    conn = get_db()

    employees = conn.execute("""
        SELECT *
        FROM employees
        WHERE site = ?
        ORDER BY name
    """, (
        site,
    )).fetchall()

    conn.close()

    return render_template(
        "attendance_select.html",
        site=site,
        month=month,
        year=year,
        day=day,
        date_name=selected_date.strftime(
            "%d %B %Y"
        ),
        day_name=selected_date.strftime(
            "%A"
        ),
        employees=employees
    )


# =========================================================
# ATTENDANCE REGISTER
# =========================================================

@app.route(
    "/attendance-register",
    methods=["POST"]
)
def attendance_register():

    try:

        site = request.form[
            "site"
        ].strip().upper()

        month = int(
            request.form["month"]
        )

        year = int(
            request.form["year"]
        )

        day = int(
            request.form["day"]
        )

    except:

        return "Invalid attendance data."

    selected_ids = request.form.getlist(
        "employee_ids"
    )

    if site not in SITES:

        return "Invalid site."

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    total_days = calendar.monthrange(
        year,
        month
    )[1]

    if day < 1 or day > total_days:

        return "Invalid date."

    if not selected_ids:

        return redirect(
            f"/attendance-select/"
            f"{site}/{month}/{year}/{day}"
        )

    try:

        selected_ids = list(
            dict.fromkeys(
                int(x)
                for x in selected_ids
            )
        )

    except:

        return "Invalid employee selection."

    selected_date = date(
        year,
        month,
        day
    )

    if selected_date > date.today():

        return "Future date attendance is locked."

    placeholders = ",".join(
        "?"
        for _ in selected_ids
    )

    conn = get_db()

    employees = conn.execute(
        f"""
        SELECT *
        FROM employees
        WHERE site = ?
        AND id IN ({placeholders})
        ORDER BY name
        """,
        [site] + selected_ids
    ).fetchall()

    attendance_date_text = (
        f"{year}-{month:02d}-{day:02d}"
    )

    records = conn.execute(
        f"""
        SELECT
            employee_id,
            status,
            ot_hours
        FROM attendance
        WHERE attendance_date = ?
        AND employee_id IN ({placeholders})
        """,
        [attendance_date_text]
        + selected_ids
    ).fetchall()

    conn.close()

    attendance_map = {

        r["employee_id"]: {

            "status": r["status"],

            "ot": r["ot_hours"]

        }

        for r in records

    }

    register_employees = []

    for employee in employees:

        saved = attendance_map.get(
            employee["id"]
        )

        register_employees.append({

            "id": employee["id"],

            "name": employee["name"],

            "employee_code":
                employee["employee_code"],

            "designation":
                employee["designation"],

            "skill":
                employee["skill"],

            "status":
                saved["status"]
                if saved else "",

            "ot":
                saved["ot"]
                if saved else ""

        })

    return render_template(
        "attendance_register.html",
        site=site,
        month=month,
        year=year,
        day=day,
        month_name=calendar.month_name[
            month
        ],
        date_name=selected_date.strftime(
            "%d %B %Y"
        ),
        day_name=selected_date.strftime(
            "%A"
        ),
        employees=register_employees
    )


# =========================================================
# SAVE DAILY ATTENDANCE
# =========================================================

@app.route(
    "/save-daily-attendance",
    methods=["POST"]
)
def save_daily_attendance():

    try:

        site = request.form[
            "site"
        ].strip().upper()

        month = int(
            request.form["month"]
        )

        year = int(
            request.form["year"]
        )

        day = int(
            request.form["day"]
        )

    except:

        return "Invalid attendance data."

    if site not in SITES:

        return "Invalid site."

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    total_days = calendar.monthrange(
        year,
        month
    )[1]

    if day < 1 or day > total_days:

        return "Invalid date."

    selected_date = date(
        year,
        month,
        day
    )

    if selected_date > date.today():

        return "Future date attendance is locked."

    employee_ids = request.form.getlist(
        "employee_ids"
    )

    conn = get_db()

    attendance_date_text = (
        f"{year}-{month:02d}-{day:02d}"
    )

    for employee_id in employee_ids:

        try:

            employee_id = int(
                employee_id
            )

        except:

            continue

        status = request.form.get(
            f"status_{employee_id}",
            ""
        )

        if status not in VALID_STATUSES:

            continue

        try:

            ot_hours = float(
                request.form.get(
                    f"ot_{employee_id}",
                    "0"
                )
            )

            ot_hours = max(
                0,
                ot_hours
            )

        except:

            ot_hours = 0

        employee_exists = conn.execute("""
            SELECT id
            FROM employees
            WHERE id = ?
            AND site = ?
        """, (
            employee_id,
            site
        )).fetchone()

        if not employee_exists:

            continue

        conn.execute("""
            INSERT INTO attendance
            (
                employee_id,
                attendance_date,
                status,
                ot_hours
            )
            VALUES (?, ?, ?, ?)

            ON CONFLICT(
                employee_id,
                attendance_date
            )

            DO UPDATE SET

                status =
                    excluded.status,

                ot_hours =
                    excluded.ot_hours
        """, (
            employee_id,
            attendance_date_text,
            status,
            ot_hours
        ))

    conn.commit()

    conn.close()

    return redirect(
        f"/attendance-dates/"
        f"{site}/{month}/{year}?saved=1"
    )


# =========================================================
# EMPLOYEE LIST
# =========================================================

@app.route(
    "/employees/<site>"
)
def employees(site):

    site = site.upper()

    search = request.args.get(
        "search",
        ""
    ).strip()

    conn = get_db()

    if site == "ALL":

        if search:

            p = f"%{search}%"

            employee_list = conn.execute("""
                SELECT *
                FROM employees
                WHERE
                    name LIKE ?
                    OR employee_code LIKE ?
                    OR designation LIKE ?
                    OR skill LIKE ?
                    OR site LIKE ?
                ORDER BY name
            """, (
                p,
                p,
                p,
                p,
                p
            )).fetchall()

        else:

            employee_list = conn.execute("""
                SELECT *
                FROM employees
                ORDER BY name
            """).fetchall()

        conn.close()

        return render_template(
            "employees.html",
            employees=employee_list,
            site="ALL",
            search=search,
            attendance_mode=False,
            manage_mode=False
        )

    if site not in SITES:

        conn.close()

        return "Invalid site."

    if search:

        p = f"%{search}%"

        employee_list = conn.execute("""
            SELECT *
            FROM employees
            WHERE site = ?
            AND (
                name LIKE ?
                OR employee_code LIKE ?
                OR designation LIKE ?
                OR skill LIKE ?
            )
            ORDER BY name
        """, (
            site,
            p,
            p,
            p,
            p
        )).fetchall()

    else:

        employee_list = conn.execute("""
            SELECT *
            FROM employees
            WHERE site = ?
            ORDER BY name
        """, (
            site,
        )).fetchall()

    conn.close()

    return render_template(
        "employees.html",
        employees=employee_list,
        site=site,
        search=search,
        attendance_mode=False,
        manage_mode=False
    )


# =========================================================
# OLD ATTENDANCE EMPLOYEE LIST
# =========================================================

@app.route(
    "/attendance-employees/<site>/<int:month>/<int:year>"
)
def attendance_employees(
    site,
    month,
    year
):

    site = site.upper()

    if site not in SITES:

        return "Invalid site."

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    search = request.args.get(
        "search",
        ""
    ).strip()

    conn = get_db()

    if search:

        p = f"%{search}%"

        employee_list = conn.execute("""
            SELECT *
            FROM employees
            WHERE site = ?
            AND (
                name LIKE ?
                OR employee_code LIKE ?
                OR designation LIKE ?
                OR skill LIKE ?
            )
            ORDER BY name
        """, (
            site,
            p,
            p,
            p,
            p
        )).fetchall()

    else:

        employee_list = conn.execute("""
            SELECT *
            FROM employees
            WHERE site = ?
            ORDER BY name
        """, (
            site,
        )).fetchall()

    conn.close()

    return render_template(
        "employees.html",
        employees=employee_list,
        site=site,
        search=search,
        attendance_mode=True,
        selected_month=month,
        selected_year=year,
        month_name=calendar.month_name[
            month
        ]
    )


# =========================================================
# ADD EMPLOYEE
# =========================================================

@app.route(
    "/add-employee",
    methods=["POST"]
)
def add_employee():

    name = request.form.get(
        "name",
        ""
    ).strip()

    employee_code = request.form.get(
        "employee_code",
        ""
    ).strip()

    designation = request.form.get(
        "designation",
        ""
    ).strip()

    skill = request.form.get(
        "skill",
        ""
    ).strip()

    site = request.form.get(
        "site",
        ""
    ).strip().upper()

    try:

        basic = max(
            0,
            float(
                request.form.get(
                    "basic",
                    0
                )
            )
        )

    except:

        basic = 0

    try:

        gross = max(
            0,
            float(
                request.form.get(
                    "gross",
                    0
                )
            )
        )

    except:

        gross = 0

    if site == "ALL":

        site = request.form.get(
            "selected_site",
            "TSM"
        ).strip().upper()

    if site not in SITES:

        return "Invalid site."

    if (
        not name
        or not employee_code
        or not designation
    ):

        return (
            "Name, Employee Code "
            "and Designation are required."
        )

    conn = get_db()

    try:

        conn.execute("""
            INSERT INTO employees
            (
                name,
                employee_code,
                designation,
                skill,
                site,
                basic,
                gross
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            name,
            employee_code,
            designation,
            skill,
            site,
            basic,
            gross
        ))

        conn.commit()

    except sqlite3.IntegrityError:

        conn.close()

        return "Employee Code already exists."

    conn.close()

    return redirect(
        "/manage-employees"
    )


# =========================================================
# EDIT EMPLOYEE
# =========================================================

@app.route(
    "/edit-employee/<int:employee_id>"
)
def edit_employee(employee_id):

    conn = get_db()

    employee = conn.execute("""
        SELECT *
        FROM employees
        WHERE id = ?
    """, (
        employee_id,
    )).fetchone()

    conn.close()

    if not employee:

        return "Employee not found."

    return render_template(
        "edit_employee.html",
        employee=employee
    )


# =========================================================
# UPDATE EMPLOYEE
# =========================================================

@app.route(
    "/update-employee/<int:employee_id>",
    methods=["POST"]
)
def update_employee(employee_id):

    name = request.form.get(
        "name",
        ""
    ).strip()

    employee_code = request.form.get(
        "employee_code",
        ""
    ).strip()

    designation = request.form.get(
        "designation",
        ""
    ).strip()

    skill = request.form.get(
        "skill",
        ""
    ).strip()

    site = request.form.get(
        "site",
        ""
    ).strip().upper()

    try:

        basic = max(
            0,
            float(
                request.form.get(
                    "basic",
                    0
                )
            )
        )

    except:

        basic = 0

    try:

        gross = max(
            0,
            float(
                request.form.get(
                    "gross",
                    0
                )
            )
        )

    except:

        gross = 0

    if site not in SITES:

        return "Invalid site."

    conn = get_db()

    try:

        conn.execute("""
            UPDATE employees
            SET
                name = ?,
                employee_code = ?,
                designation = ?,
                skill = ?,
                site = ?,
                basic = ?,
                gross = ?
            WHERE id = ?
        """, (
            name,
            employee_code,
            designation,
            skill,
            site,
            basic,
            gross,
            employee_id
        ))

        conn.commit()

    except sqlite3.IntegrityError:

        conn.close()

        return "Employee Code already exists."

    conn.close()

    return redirect(
        "/manage-employees"
    )


# =========================================================
# IMPORT EMPLOYEES FROM EXCEL
# =========================================================

@app.route(
    "/import-employees",
    methods=["POST"]
)
def import_employees():

    site = request.form.get(
        "site",
        "TSM"
    ).strip().upper()

    if site == "ALL":

        site = "TSM"

    if site not in SITES:

        site = "TSM"

    file = request.files.get(
        "excel_file"
    )

    if not file or not file.filename:

        return redirect(
            "/manage-employees"
        )

    if not file.filename.lower().endswith(
        ".xlsx"
    ):

        return redirect(
            "/manage-employees"
        )

    temp_file = (
        "temp_employee_import.xlsx"
    )

    file.save(temp_file)

    conn = get_db()

    try:

        workbook = load_workbook(
            temp_file,
            data_only=True
        )

        sheet = workbook.active

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            name = (
                row[0]
                if len(row) > 0
                else None
            )

            code = (
                row[1]
                if len(row) > 1
                else None
            )

            designation = (
                row[2]
                if len(row) > 2
                else None
            )

            skill = (
                row[3]
                if len(row) > 3
                else ""
            )

            excel_site = (
                row[4]
                if len(row) > 4
                else site
            )

            basic = (
                row[5]
                if len(row) > 5
                else 0
            )

            gross = (
                row[6]
                if len(row) > 6
                else 0
            )

            if (
                not name
                or not code
                or not designation
            ):

                continue

            name = str(name).strip()

            code = str(code).strip()

            designation = str(
                designation
            ).strip()

            skill = (
                str(skill).strip()
                if skill
                else ""
            )

            excel_site = (
                str(excel_site)
                .strip()
                .upper()
                if excel_site
                else site
            )

            if excel_site not in SITES:

                excel_site = site

            try:

                basic = max(
                    0,
                    float(basic or 0)
                )

            except:

                basic = 0

            try:

                gross = max(
                    0,
                    float(gross or 0)
                )

            except:

                gross = 0

            existing = conn.execute("""
                SELECT id
                FROM employees
                WHERE employee_code = ?
            """, (
                code,
            )).fetchone()

            if existing:

                conn.execute("""
                    UPDATE employees
                    SET
                        name = ?,
                        designation = ?,
                        skill = ?,
                        site = ?,
                        basic = ?,
                        gross = ?
                    WHERE employee_code = ?
                """, (
                    name,
                    designation,
                    skill,
                    excel_site,
                    basic,
                    gross,
                    code
                ))

            else:

                conn.execute("""
                    INSERT INTO employees
                    (
                        name,
                        employee_code,
                        designation,
                        skill,
                        site,
                        basic,
                        gross
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    name,
                    code,
                    designation,
                    skill,
                    excel_site,
                    basic,
                    gross
                ))

        conn.commit()

    except Exception as e:

        conn.rollback()

        conn.close()

        if os.path.exists(temp_file):

            os.remove(temp_file)

        return (
            f"Excel import error: {e}"
        )

    conn.close()

    if os.path.exists(temp_file):

        os.remove(temp_file)

    return redirect(
        "/manage-employees"
    )


# =========================================================
# OLD ATTENDANCE
# =========================================================

@app.route(
    "/attendance/<int:employee_id>/<int:month>/<int:year>"
)
def attendance(
    employee_id,
    month,
    year
):

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    today = date.today()

    conn = get_db()

    employee = conn.execute("""
        SELECT *
        FROM employees
        WHERE id = ?
    """, (
        employee_id,
    )).fetchone()

    if not employee:

        conn.close()

        return "Employee not found."

    site_employees = conn.execute("""
        SELECT *
        FROM employees
        WHERE site = ?
        ORDER BY name
    """, (
        employee["site"],
    )).fetchall()

    ids = [
        e["id"]
        for e in site_employees
    ]

    current_index = ids.index(
        employee_id
    )

    previous_employee = (

        site_employees[
            current_index - 1
        ]

        if current_index > 0

        else None

    )

    next_employee = (

        site_employees[
            current_index + 1
        ]

        if current_index
        < len(site_employees) - 1

        else None

    )

    records = conn.execute("""
        SELECT
            attendance_date,
            status,
            ot_hours
        FROM attendance
        WHERE employee_id = ?
        AND attendance_date LIKE ?
    """, (
        employee_id,
        f"{year}-{month:02d}-%"
    )).fetchall()

    conn.close()

    attendance_data = {}

    for record in records:

        day = int(
            record[
                "attendance_date"
            ].split("-")[2]
        )

        attendance_data[day] = {

            "status":
                record["status"],

            "ot":
                record["ot_hours"]

        }

    total_days = calendar.monthrange(
        year,
        month
    )[1]

    calendar_days = []

    for d in range(
        1,
        total_days + 1
    ):

        current = date(
            year,
            month,
            d
        )

        saved = attendance_data.get(
            d
        )

        calendar_days.append({

            "date": d,

            "day":
                current.strftime("%a"),

            "full_date":
                current.strftime("%d-%m-%Y"),

            "status":
                saved["status"]
                if saved else "",

            "ot":
                saved["ot"]
                if saved else "",

            "future":
                current > today

        })

    present_days = 0
    absent_days = 0
    wo_days = 0
    nh_days = 0
    pl_days = 0
    total_ot = 0

    for item in calendar_days:

        status = item["status"]

        if status == "P":

            present_days += 1

        elif status == "A":

            absent_days += 1

        elif status == "W/O":

            wo_days += 1

        elif status == "NH":

            nh_days += 1

        elif status == "PL":

            pl_days += 1

        try:

            total_ot += float(
                item["ot"] or 0
            )

        except:

            pass

    return render_template(
        "attendance.html",
        employee=employee,
        calendar_days=calendar_days,
        month=month,
        year=year,
        month_name=calendar.month_name[
            month
        ],
        previous_employee=previous_employee,
        next_employee=next_employee,
        today=today.strftime(
            "%Y-%m-%d"
        ),
        present_days=present_days,
        absent_days=absent_days,
        wo_days=wo_days,
        nh_days=nh_days,
        fl_days=pl_days,
        total_ot=total_ot,
        basic=employee["basic"],
        gross=employee["gross"]
    )


# =========================================================
# OLD SAVE ATTENDANCE
# =========================================================

@app.route(
    "/save-attendance",
    methods=["POST"]
)
def save_attendance():

    employee_id = int(
        request.form["employee_id"]
    )

    month = int(
        request.form["month"]
    )

    year = int(
        request.form["year"]
    )

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    today = date.today()

    conn = get_db()

    total_days = calendar.monthrange(
        year,
        month
    )[1]

    for d in range(
        1,
        total_days + 1
    ):

        attendance_date = date(
            year,
            month,
            d
        )

        if attendance_date > today:

            continue

        status = request.form.get(
            f"status_{d}",
            ""
        )

        if status not in VALID_STATUSES:

            continue

        try:

            ot_hours = max(
                0,
                float(
                    request.form.get(
                        f"ot_{d}",
                        0
                    )
                )
            )

        except:

            ot_hours = 0

        text_date = (
            f"{year}-{month:02d}-{d:02d}"
        )

        conn.execute("""
            INSERT INTO attendance
            (
                employee_id,
                attendance_date,
                status,
                ot_hours
            )
            VALUES (?, ?, ?, ?)

            ON CONFLICT(
                employee_id,
                attendance_date
            )

            DO UPDATE SET

                status =
                    excluded.status,

                ot_hours =
                    excluded.ot_hours
        """, (
            employee_id,
            text_date,
            status,
            ot_hours
        ))

    conn.commit()

    conn.close()

    return redirect(
        f"/attendance/"
        f"{employee_id}/{month}/{year}"
        f"?saved=1"
    )


# =========================================================
# SHOW ATTENDANCE
# =========================================================

@app.route("/show-attendance")
def show_attendance():

    return render_template(
        "show_attendance_site.html",
        sites=SITES
    )


@app.route(
    "/show-attendance/<site>"
)
def show_attendance_month(site):

    site = site.upper()

    if site not in SITES:

        return "Invalid site."

    return render_template(
        "show_attendance_month.html",
        site=site,
        months=get_fy_months()
    )


# =========================================================
# SALARY CALCULATION
# =========================================================

def calculate_employee_salary(
    basic_rate,
    gross_rate,
    present_days,
    ot_hours,
    profession_tax=0,
    advance_deduction=0,
    other_deduction=0
):

    basic_rate = float(
        basic_rate or 0
    )

    gross_rate = float(
        gross_rate or 0
    )

    present_days = int(
        present_days or 0
    )

    ot_hours = float(
        ot_hours or 0
    )

    profession_tax = max(
        0,
        float(
            profession_tax or 0
        )
    )

    advance_deduction = max(
        0,
        float(
            advance_deduction or 0
        )
    )

    other_deduction = max(
        0,
        float(
            other_deduction or 0
        )
    )

    # -----------------------------------------------------
    # BASIC
    # -----------------------------------------------------

    basic_daily = (
        basic_rate / 26
    )

    basic_payable = (
        basic_daily * present_days
    )

    # -----------------------------------------------------
    # GROSS
    # -----------------------------------------------------

    gross_daily = (
        gross_rate / 26
    )

    gross_payable = (
        gross_daily * present_days
    )

    # -----------------------------------------------------
    # EPF / ESIC
    # -----------------------------------------------------

    epf = (
        basic_payable
        * EPF_RATE
        / 100
    )

    esic = (
        gross_payable
        * ESIC_RATE
        / 100
    )

    # -----------------------------------------------------
    # FIRST PAYMENT BEFORE DEDUCTIONS
    # -----------------------------------------------------

    first_payment_before_deduction = max(
        0,
        basic_payable
        - epf
        - esic
    )

    # -----------------------------------------------------
    # OTHER DEDUCTIONS
    # -----------------------------------------------------

    total_deductions = (
        profession_tax
        + advance_deduction
        + other_deduction
    )

    # -----------------------------------------------------
    # FINAL FIRST PAYMENT
    # -----------------------------------------------------

    first_payment = max(
        0,
        first_payment_before_deduction
        - total_deductions
    )

    # -----------------------------------------------------
    # OT
    # -----------------------------------------------------

    basic_hourly = (
        basic_daily / 8
    )

    ot_hourly = (
        basic_hourly
        * OT_MULTIPLIER
    )

    ot_payable = (
        ot_hours
        * ot_hourly
    )

    # -----------------------------------------------------
    # SECOND PAYMENT
    # -----------------------------------------------------

    second_payment = max(
        0,
        gross_payable
        - basic_payable
        + ot_payable
    )

    # -----------------------------------------------------
    # TOTAL
    # -----------------------------------------------------

    total_payment = (
        first_payment
        + second_payment
    )

    return {

        "basic_rate":
            round(
                basic_rate,
                2
            ),

        "gross_rate":
            round(
                gross_rate,
                2
            ),

        "basic_payable":
            round(
                basic_payable,
                2
            ),

        "gross_payable":
            round(
                gross_payable,
                2
            ),

        "epf":
            round(
                epf,
                2
            ),

        "esic":
            round(
                esic,
                2
            ),

        "first_payment_before_deduction":
            round(
                first_payment_before_deduction,
                2
            ),

        "profession_tax":
            round(
                profession_tax,
                2
            ),

        "advance_deduction":
            round(
                advance_deduction,
                2
            ),

        "other_deduction":
            round(
                other_deduction,
                2
            ),

        "total_deductions":
            round(
                total_deductions,
                2
            ),

        "first_payment":
            round(
                first_payment,
                2
            ),

        "basic_hourly":
            round(
                basic_hourly,
                2
            ),

        "ot_payable":
            round(
                ot_payable,
                2
            ),

        "second_payment":
            round(
                second_payment,
                2
            ),

        "total_payment":
            round(
                total_payment,
                2
            )

    }


# =========================================================
# SAVE SALARY DEDUCTIONS
# =========================================================

@app.route(
    "/save-salary-deductions",
    methods=["POST"]
)
def save_salary_deductions():

    try:

        employee_id = int(
            request.form[
                "employee_id"
            ]
        )

        month = int(
            request.form[
                "month"
            ]
        )

        year = int(
            request.form[
                "year"
            ]
        )

        profession_tax = max(
            0,
            float(
                request.form.get(
                    "profession_tax",
                    0
                ) or 0
            )
        )

        advance_deduction = max(
            0,
            float(
                request.form.get(
                    "advance_deduction",
                    0
                ) or 0
            )
        )

        other_deduction = max(
            0,
            float(
                request.form.get(
                    "other_deduction",
                    0
                ) or 0
            )
        )

    except:

        return "Invalid deduction data."

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid salary month."

    conn = get_db()

    conn.execute("""
        INSERT INTO salary_deductions
        (
            employee_id,
            month,
            year,
            profession_tax,
            advance_deduction,
            other_deduction
        )
        VALUES (?, ?, ?, ?, ?, ?)

        ON CONFLICT(
            employee_id,
            month,
            year
        )

        DO UPDATE SET

            profession_tax =
                excluded.profession_tax,

            advance_deduction =
                excluded.advance_deduction,

            other_deduction =
                excluded.other_deduction

    """, (
        employee_id,
        month,
        year,
        profession_tax,
        advance_deduction,
        other_deduction
    ))

    conn.commit()

    conn.close()

    site = request.form.get(
        "site",
        "TSM"
    ).strip().upper()

    if site not in SITES:

        site = "TSM"

    return redirect(
        f"/show-attendance/"
        f"{site}/{month}/{year}"
    )


# =========================================================
# BUILD REPORT DATA
# =========================================================

def get_attendance_report_data(
    site,
    month,
    year
):

    conn = get_db()

    employees_db = conn.execute("""
        SELECT *
        FROM employees
        WHERE site = ?
        ORDER BY name
    """, (
        site,
    )).fetchall()

    total_days = calendar.monthrange(
        year,
        month
    )[1]

    calendar_days = list(
        range(
            1,
            total_days + 1
        )
    )

    totals = {

        "present_days": 0,

        "ot_hours": 0,

        "basic_payable": 0,

        "gross_payable": 0,

        "epf": 0,

        "esic": 0,

        "first_payment_before_deduction": 0,

        "profession_tax": 0,

        "advance_deduction": 0,

        "other_deduction": 0,

        "total_deductions": 0,

        "first_payment": 0,

        "ot_payable": 0,

        "second_payment": 0,

        "total_payment": 0

    }

    report_employees = []

    for employee in employees_db:

        records = conn.execute("""
            SELECT
                attendance_date,
                status,
                ot_hours
            FROM attendance
            WHERE employee_id = ?
            AND attendance_date LIKE ?
        """, (
            employee["id"],
            f"{year}-{month:02d}-%"
        )).fetchall()

        attendance_map = {}

        for record in records:

            d = int(
                record[
                    "attendance_date"
                ].split("-")[2]
            )

            attendance_map[d] = {

                "status":
                    record["status"],

                "ot":
                    float(
                        record["ot_hours"]
                        or 0
                    )

            }

        deduction = conn.execute("""
            SELECT
                profession_tax,
                advance_deduction,
                other_deduction
            FROM salary_deductions
            WHERE employee_id = ?
            AND month = ?
            AND year = ?
        """, (
            employee["id"],
            month,
            year
        )).fetchone()

        profession_tax = (

            deduction[
                "profession_tax"
            ]

            if deduction

            else 0

        )

        advance_deduction = (

            deduction[
                "advance_deduction"
            ]

            if deduction

            else 0

        )

        other_deduction = (

            deduction[
                "other_deduction"
            ]

            if deduction

            else 0

        )

        days = []

        daily_attendance = []

        present_days = 0

        actual_present_days = 0

        fl_days = 0

        nh_days = 0

        ot_hours = 0

        for d in calendar_days:

            record = attendance_map.get(
                d
            )

            status = (
                record["status"]
                if record
                else ""
            )

            ot = (
                record["ot"]
                if record
                else 0
            )

            days.append(status)

            daily_attendance.append({
                "status": status,
                "ot": float(ot or 0)
            })

            if status == "P":

                actual_present_days += 1

            if status == "FL":

                fl_days += 1

            if status == "NH":

                nh_days += 1

            if status in ("P", "FL", "NH"):

                present_days += 1

            ot_hours += float(
                ot or 0
            )

        salary = calculate_employee_salary(

            employee["basic"],

            employee["gross"],

            present_days,

            ot_hours,

            profession_tax,

            advance_deduction,

            other_deduction

        )

        report_employee = {

            "id":
                employee["id"],

            "name":
                employee["name"],

            "employee_code":
                employee["employee_code"],

            "designation":
                employee["designation"],

            "skill":
                employee["skill"],

            "site":
                employee["site"],

            "days":
                days,

            "present_days":
                present_days,

            "daily_attendance":
                daily_attendance,

            "actual_present_days":
                actual_present_days,

            "fl_days":
                fl_days,

            "nh_days":
                nh_days,

            "ot_hours":
                round(
                    ot_hours,
                    2
                ),

            "basic_rate":
                salary["basic_rate"],

            "gross_rate":
                salary["gross_rate"],

            "basic_payable":
                salary["basic_payable"],

            "gross_payable":
                salary["gross_payable"],

            "epf":
                salary["epf"],

            "esic":
                salary["esic"],

            "first_payment_before_deduction":
                salary[
                    "first_payment_before_deduction"
                ],

            "profession_tax":
                salary[
                    "profession_tax"
                ],

            "advance_deduction":
                salary[
                    "advance_deduction"
                ],

            "other_deduction":
                salary[
                    "other_deduction"
                ],

            "total_deductions":
                salary[
                    "total_deductions"
                ],

            "first_payment":
                salary[
                    "first_payment"
                ],

            "ot_payable":
                salary[
                    "ot_payable"
                ],

            "second_payment":
                salary[
                    "second_payment"
                ],

            "total_payment":
                salary[
                    "total_payment"
                ]

        }

        report_employees.append(
            report_employee
        )

        for key in totals:

            if key == "present_days":

                totals[key] += present_days

            elif key == "ot_hours":

                totals[key] += ot_hours

            elif key in salary:

                totals[key] += salary[key]

    conn.close()

    for key in totals:

        totals[key] = round(
            totals[key],
            2
        )

    return (
        report_employees,
        calendar_days,
        totals
    )


# =========================================================
# SHOW ATTENDANCE REPORT
# =========================================================

@app.route(
    "/show-attendance/"
    "<site>/<int:month>/<int:year>"
)
def show_attendance_report(
    site,
    month,
    year
):

    site = site.upper()

    if site not in SITES:

        return "Invalid site."

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    (
        report_employees,
        calendar_days,
        totals
    ) = get_attendance_report_data(
        site,
        month,
        year
    )

    return render_template(
        "show_attendance_report.html",

        site=site,

        month=month,

        year=year,

        month_name=calendar.month_name[
            month
        ],

        calendar_days=calendar_days,

        employees=report_employees,

        totals=totals
    )


# =========================================================
# EXPORT EXCEL
# =========================================================

@app.route(
    "/export-attendance-excel/"
    "<site>/<int:month>/<int:year>"
)
def export_attendance_excel(
    site,
    month,
    year
):

    site = site.upper()

    if site not in SITES:

        return "Invalid site."

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    (
        employees,
        calendar_days,
        totals
    ) = get_attendance_report_data(
        site,
        month,
        year
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = (
        f"{site} {calendar.month_name[month]}"
    )

    headers = [

        "Employee Name",

        "Emp Code",

        "Designation",

        "Skill"

    ]

    for day in calendar_days:

        headers.append(
            str(day)
        )

    headers += [

        "Present",

        "OT Hours",

        "Basic Rate",

        "Gross Rate",

        "Basic Payable",

        "Gross Payable",

        "EPF",

        "ESIC",

        "1st Payment Before Deduction",

        "Profession Tax",

        "Advance Deduction",

        "Other Deduction",

        "Total Deduction",

        "1st Payment",

        "OT Payable",

        "2nd Payment",

        "Total Payment"

    ]

    sheet.append(headers)

    # -----------------------------------------------------
    # HEADER STYLE
    # -----------------------------------------------------

    header_fill = PatternFill(
        "solid",
        fgColor="12263D"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_border = Border(
        bottom=Side(
            style="thin",
            color="999999"
        )
    )

    for cell in sheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center"
        )

        cell.border = thin_border

    # -----------------------------------------------------
    # DATA
    # -----------------------------------------------------

    for employee in employees:

        row = [

            employee["name"],

            employee["employee_code"],

            employee["designation"],

            employee["skill"]

        ]

        row += employee["days"]

        row += [

            employee["present_days"],

            employee["ot_hours"],

            employee["basic_rate"],

            employee["gross_rate"],

            employee["basic_payable"],

            employee["gross_payable"],

            employee["epf"],

            employee["esic"],

            employee[
                "first_payment_before_deduction"
            ],

            employee["profession_tax"],

            employee["advance_deduction"],

            employee["other_deduction"],

            employee["total_deductions"],

            employee["first_payment"],

            employee["ot_payable"],

            employee["second_payment"],

            employee["total_payment"]

        ]

        sheet.append(row)

    # -----------------------------------------------------
    # TOTAL ROW
    # -----------------------------------------------------

    total_row = [

        "TOTAL",

        "",

        "",

        ""

    ]

    total_row += [
        ""
        for _ in calendar_days
    ]

    total_row += [

        totals["present_days"],

        totals["ot_hours"],

        "",

        "",

        totals["basic_payable"],

        totals["gross_payable"],

        totals["epf"],

        totals["esic"],

        totals[
            "first_payment_before_deduction"
        ],

        totals["profession_tax"],

        totals["advance_deduction"],

        totals["other_deduction"],

        totals["total_deductions"],

        totals["first_payment"],

        totals["ot_payable"],

        totals["second_payment"],

        totals["total_payment"]

    ]

    sheet.append(total_row)

    total_index = sheet.max_row

    for cell in sheet[total_index]:

        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )

    # -----------------------------------------------------
    # COLUMN WIDTHS
    # -----------------------------------------------------

    for column in sheet.columns:

        max_length = 0

        column_letter = (
            column[0].column_letter
        )

        for cell in column:

            try:

                length = len(
                    str(cell.value)
                )

                max_length = max(
                    max_length,
                    length
                )

            except:

                pass

        sheet.column_dimensions[
            column_letter
        ].width = min(
            max(max_length + 2, 12),
            30
        )

    sheet.freeze_panes = "A2"

    # -----------------------------------------------------
    # OUTPUT
    # -----------------------------------------------------

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    filename = (
        f"{site}_Attendance_"
        f"{calendar.month_name[month]}_"
        f"{year}.xlsx"
    )

    return send_file(

        output,

        as_attachment=True,

        download_name=filename,

        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )

    )


# =========================================================
# SALARY SLIP PDF
# =========================================================

def create_salary_slip_pdf(
    site,
    month,
    year,
    employees
):

    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=28,
        leftMargin=28,
        topMargin=25,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()

    normal = ParagraphStyle(
        "SlipNormal",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        alignment=TA_LEFT
    )

    center = ParagraphStyle(
        "SlipCenter",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        alignment=TA_CENTER
    )

    center_bold = ParagraphStyle(
        "SlipCenterBold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        alignment=TA_CENTER
    )

    right = ParagraphStyle(
        "SlipRight",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        alignment=TA_RIGHT
    )

    right_bold = ParagraphStyle(
        "SlipRightBold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        alignment=TA_RIGHT
    )

    title = ParagraphStyle(
        "SlipTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=16,
        alignment=TA_CENTER
    )

    company = ParagraphStyle(
        "Company",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=15,
        alignment=TA_CENTER
    )

    address = ParagraphStyle(
        "Address",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        alignment=TA_CENTER
    )

    story = []

    month_name = calendar.month_name[month]

    # -----------------------------------------------------
    # TWO SLIPS PER A4 PAGE
    # -----------------------------------------------------

    for index, employee in enumerate(employees):

        # =================================================
        # COMPANY HEADER
        # =================================================

        header_data = [

            [
                Paragraph(
                    "FORM XV",
                    center_bold
                )
            ],

            [
                Paragraph(
                    "(Rule 77 (2) (h))",
                    center
                )
            ],

            [
                Paragraph(
                    "SD CONSTRUCTION",
                    company
                )
            ],

            [
                Paragraph(
                    "Plot No- 479/3325, Mouza- Trijanga, "
                    "Darnagadi, Jaipur Road, Jajpur, Odisha, 755026",
                    address
                )
            ],

            [
                Paragraph(
                    f"Wage Slip for the month of "
                    f"{month_name}'{str(year)[-2:]}",
                    center_bold
                )
            ]

        ]

        header_table = Table(
            header_data,
            colWidths=[490]
        )

        header_table.setStyle(
            TableStyle([

                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER"
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, 0),
                    2
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    2
                )

            ])
        )

        story.append(header_table)

        story.append(
            Spacer(1, 5)
        )

        # =================================================
        # EMPLOYEE DETAILS
        # =================================================

        employee_details = [

            [
                Paragraph("<b>Name</b>", normal),
                Paragraph(
                    f": {employee['name']}",
                    normal
                ),

                Paragraph(
                    "<b>Employee Code</b>",
                    normal
                ),

                Paragraph(
                    f": {employee['employee_code']}",
                    normal
                )
            ],

            [
                Paragraph(
                    "<b>Designation</b>",
                    normal
                ),

                Paragraph(
                    f": {employee['designation']}",
                    normal
                ),

                Paragraph(
                    "<b>Present Days</b>",
                    normal
                ),

                Paragraph(
                    f": {employee['actual_present_days']}",
                    normal
                )
            ],

            [
                Paragraph(
                    "<b>FL</b>",
                    normal
                ),

                Paragraph(
                    f": {employee['fl_days']}",
                    normal
                ),

                Paragraph(
                    "<b>NH</b>",
                    normal
                ),

                Paragraph(
                    f": {employee['nh_days']}",
                    normal
                )
            ],

            [
                Paragraph(
                    "<b>Employee Branch</b>",
                    normal
                ),

                Paragraph(
                    f": {site}",
                    normal
                ),

                Paragraph(
                    "<b>OT Hours</b>",
                    normal
                ),

                Paragraph(
                    f": {employee['ot_hours']}",
                    normal
                )
            ]

        ]

        employee_table = Table(
            employee_details,
            colWidths=[
                80,
                190,
                90,
                130
            ]
        )

        employee_table.setStyle(
            TableStyle([

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    4
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    4
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    3
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    3
                )

            ])
        )

        story.append(
            employee_table
        )

        story.append(
            Spacer(1, 5)
        )

        # =================================================
        # CALCULATIONS
        # =================================================

        basic_salary = float(
            employee["basic_payable"] or 0
        )

        ot_payable = float(
            employee["ot_payable"] or 0
        )

        epf = float(
            employee["epf"] or 0
        )

        esic = float(
            employee["esic"] or 0
        )

        profession_tax = float(
            employee["profession_tax"] or 0
        )

        advance_deduction = float(
            employee["advance_deduction"] or 0
        )

        other_deduction = float(
            employee["other_deduction"] or 0
        )

        other_allowance = float(
            employee["second_payment"] or 0
        )

        # Total earning = Basic + HRA + Other Allowance
        # + Incentive + OT
        #
        # IMPORTANT:
        # Other Allowance earning side remains ZERO.
        total_earnings = (
            basic_salary
            + ot_payable
        )

        total_deductions = (
            epf
            + esic
            + profession_tax
            + advance_deduction
            + other_deduction
        )

        net_pay = max(
            0,
            total_earnings - total_deductions
        )

        # =================================================
        # EARNINGS / DEDUCTIONS
        # =================================================

        earnings_data = [

            [
                Paragraph(
                    "<b>Earnings</b>",
                    center_bold
                ),

                Paragraph(
                    "<b>Amount (INR)</b>",
                    center_bold
                ),

                Paragraph(
                    "<b>Deductions</b>",
                    center_bold
                ),

                Paragraph(
                    "<b>Amount (INR)</b>",
                    center_bold
                )
            ],

            [
                Paragraph(
                    "Basic Salary",
                    normal
                ),

                Paragraph(
                    f"{basic_salary:.2f}",
                    right
                ),

                Paragraph(
                    "EPF",
                    normal
                ),

                Paragraph(
                    f"{epf:.2f}",
                    right
                )
            ],

            [
                Paragraph(
                    "HRA",
                    normal
                ),

                Paragraph(
                    "0.00",
                    right
                ),

                Paragraph(
                    "ESIC",
                    normal
                ),

                Paragraph(
                    f"{esic:.2f}",
                    right
                )
            ],

            [
                Paragraph(
                    "Other Allowances",
                    normal
                ),

                Paragraph(
                    "0.00",
                    right
                ),

                Paragraph(
                    "Prof. Tax",
                    normal
                ),

                Paragraph(
                    f"{profession_tax:.2f}",
                    right
                )
            ],

            [
                Paragraph(
                    "Incentive",
                    normal
                ),

                Paragraph(
                    "0.00",
                    right
                ),

                Paragraph(
                    "Advance",
                    normal
                ),

                Paragraph(
                    f"{advance_deduction:.2f}",
                    right
                )
            ],

            [
                Paragraph(
                    "OT",
                    normal
                ),

                Paragraph(
                    f"{ot_payable:.2f}",
                    right
                ),

                Paragraph(
                    "Other Deduction",
                    normal
                ),

                Paragraph(
                    f"{other_deduction:.2f}",
                    right
                )
            ],

            [
                Paragraph(
                    "<b>Total Earnings (INR)</b>",
                    normal
                ),

                Paragraph(
                    f"<b>{total_earnings:.2f}</b>",
                    right_bold
                ),

                Paragraph(
                    "<b>Total Deductions (INR)</b>",
                    normal
                ),

                Paragraph(
                    f"<b>{total_deductions:.2f}</b>",
                    right_bold
                )
            ]

        ]

        earnings_table = Table(
            earnings_data,
            colWidths=[
                155,
                90,
                155,
                90
            ]
        )

        earnings_table.setStyle(
            TableStyle([

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#E8EEF5")
                ),

                (
                    "BACKGROUND",
                    (0, -1),
                    (-1, -1),
                    colors.HexColor("#F2F2F2")
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),

                (
                    "ALIGN",
                    (1, 1),
                    (1, -1),
                    "RIGHT"
                ),

                (
                    "ALIGN",
                    (3, 1),
                    (3, -1),
                    "RIGHT"
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    3
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    3
                )

            ])
        )

        story.append(
            earnings_table
        )

        story.append(
            Spacer(1, 4)
        )

        # =================================================
        # NET PAY
        # =================================================

        net_data = [

            [
                Paragraph(
                    "<b>Net Pay for This Month "
                    "(Total Earnings - Total Deductions):</b>",
                    normal
                ),

                Paragraph(
                    f"<b>{net_pay:.2f}</b>",
                    right_bold
                )
            ],

            [
                Paragraph(
                    "<b>Other Allowances for This Month:</b>",
                    normal
                ),

                Paragraph(
                    f"<b>{other_allowance:.2f}</b>",
                    right_bold
                )
            ]

        ]

        net_table = Table(
            net_data,
            colWidths=[
                350,
                140
            ]
        )

        net_table.setStyle(
            TableStyle([

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.6,
                    colors.grey
                ),

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#F2F2F2")
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    4
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4
                )

            ])
        )

        story.append(
            net_table
        )

        story.append(
            Spacer(1, 4)
        )

        story.append(
            Paragraph(
                "This is a system generated payslip. "
                "Hence does not require signature.",
                center
            )
        )

        # =================================================
        # TWO SLIPS PER PAGE
        # =================================================

        if index % 2 == 1:

            if index != len(employees) - 1:
                story.append(
                    PageBreak()
                )

        else:

            if index != len(employees) - 1:

                story.append(
                    Spacer(1, 10)
                )

                divider = Table(
                    [[""]],
                    colWidths=[490],
                    rowHeights=[1]
                )

                divider.setStyle(
                    TableStyle([
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, -1),
                            colors.grey
                        )
                    ])
                )

                story.append(
                    divider
                )

                story.append(
                    Spacer(1, 10)
                )

    doc.build(story)

    output.seek(0)

    return output


# =========================================================
# EXPORT SALARY SLIPS PDF
# =========================================================

@app.route(
    "/export-salary-slips/"
    "<site>/<int:month>/<int:year>"
)
def export_salary_slips(
    site,
    month,
    year
):

    site = site.upper()

    if site not in SITES:
        return "Invalid site."

    if not valid_financial_year(
        month,
        year
    ):
        return "Invalid salary month."

    (
        employees,
        calendar_days,
        totals
    ) = get_attendance_report_data(
        site,
        month,
        year
    )

    if not employees:
        return "No employees found."

    pdf = create_salary_slip_pdf(
        site,
        month,
        year,
        employees
    )

    filename = (
        f"{site}_Salary_Slips_"
        f"{calendar.month_name[month]}_"
        f"{year}.pdf"
    )

    return send_file(
        pdf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


# =========================================================
# ATTENDANCE PDF EXPORT
# =========================================================

def create_attendance_report_pdf(
    site,
    month,
    year,
    employees,
    calendar_days,
    totals
):

    output = BytesIO()

    # -----------------------------------------------------
    # LANDSCAPE A4
    # -----------------------------------------------------

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=12,
        leftMargin=12,
        topMargin=15,
        bottomMargin=15
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "AttendancePDFTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=15,
        alignment=TA_CENTER
    )

    center_style = ParagraphStyle(
        "AttendanceCenter",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.5,
        leading=6,
        alignment=TA_CENTER
    )

    header_style = ParagraphStyle(
        "AttendanceHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=5.2,
        leading=5.8,
        alignment=TA_CENTER
    )

    story = []

    month_name = calendar.month_name[month]

    # =====================================================
    # REPORT HEADER
    # =====================================================

    story.append(
        Paragraph(
            "SD CONSTRUCTION",
            title_style
        )
    )

    story.append(
        Paragraph(
            "Attendance & Salary Report",
            center_style
        )
    )

    story.append(
        Paragraph(
            f"SITE: {site} | {month_name} {year}",
            center_style
        )
    )

    story.append(
        Spacer(1, 6)
    )

    # =====================================================
    # TABLE HEADER
    # =====================================================

    table_header = [

        Paragraph(
            "Employee",
            header_style
        ),

        Paragraph(
            "Code",
            header_style
        ),

        Paragraph(
            "Designation",
            header_style
        ),

        Paragraph(
            "Skill",
            header_style
        )

    ]

    # Attendance dates

    for day in calendar_days:

        table_header.append(
            Paragraph(
                str(day),
                header_style
            )
        )

    # Salary columns

    table_header.extend([

        Paragraph(
            "P",
            header_style
        ),

        Paragraph(
            "OT",
            header_style
        ),

        Paragraph(
            "Basic",
            header_style
        ),

        Paragraph(
            "Gross",
            header_style
        ),

        Paragraph(
            "EPF",
            header_style
        ),

        Paragraph(
            "ESIC",
            header_style
        ),

        Paragraph(
            "Prof. Tax",
            header_style
        ),

        Paragraph(
            "Advance",
            header_style
        ),

        Paragraph(
            "Other Ded.",
            header_style
        ),

        Paragraph(
            "Net Pay",
            header_style
        ),

        Paragraph(
            "Other Allow.",
            header_style
        ),

        Paragraph(
            "Total",
            header_style
        )

    ])

    table_data = [
        table_header
    ]

    # =====================================================
    # EMPLOYEE ROWS
    # =====================================================

    for employee in employees:

        row = [

            Paragraph(
                str(employee["name"]),
                center_style
            ),

            Paragraph(
                str(employee["employee_code"]),
                center_style
            ),

            Paragraph(
                str(employee["designation"]),
                center_style
            ),

            Paragraph(
                str(employee["skill"]),
                center_style
            )

        ]

        # -------------------------------------------------
        # ATTENDANCE
        # -------------------------------------------------

        for status in employee["days"]:

            row.append(
                Paragraph(
                    str(status),
                    center_style
                )
            )

        # -------------------------------------------------
        # SALARY
        # -------------------------------------------------

        row.extend([

            Paragraph(
                str(employee["present_days"]),
                center_style
            ),

            Paragraph(
                f"{employee['ot_hours']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['basic_payable']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['gross_payable']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['epf']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['esic']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['profession_tax']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['advance_deduction']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['other_deduction']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['first_payment']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['second_payment']:.2f}",
                center_style
            ),

            Paragraph(
                f"{employee['total_payment']:.2f}",
                center_style
            )

        ])

        table_data.append(row)

    # =====================================================
    # TOTAL ROW
    # =====================================================

    total_row = [

        Paragraph(
            "<b>TOTAL</b>",
            header_style
        ),

        "",
        "",
        ""

    ]

    # Empty attendance cells

    total_row.extend(
        [
            ""
            for _ in calendar_days
        ]
    )

    # Totals

    total_row.extend([

        Paragraph(
            f"<b>{totals['present_days']:.0f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['ot_hours']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['basic_payable']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['gross_payable']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['epf']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['esic']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['profession_tax']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['advance_deduction']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['other_deduction']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['first_payment']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['second_payment']:.2f}</b>",
            header_style
        ),

        Paragraph(
            f"<b>{totals['total_payment']:.2f}</b>",
            header_style
        )

    ])

    table_data.append(total_row)

    # =====================================================
    # COLUMN WIDTHS
    # =====================================================

    col_widths = [

        72,   # Employee
        42,   # Code
        58,   # Designation
        42    # Skill

    ]

    # Attendance days

    col_widths += [

        14
        for _ in calendar_days

    ]

    # Salary columns

    col_widths += [

        24,   # Present
        27,   # OT
        42,   # Basic
        42,   # Gross
        38,   # EPF
        38,   # ESIC
        42,   # Profession Tax
        44,   # Advance
        44,   # Other Deduction
        45,   # Net Pay
        45,   # Other Allowance
        45    # Total

    ]

    # =====================================================
    # CREATE TABLE
    # =====================================================

    attendance_table = Table(
        table_data,
        colWidths=col_widths,
        repeatRows=1,
        splitByRow=1
    )

    # =====================================================
    # TABLE STYLE
    # =====================================================

    attendance_table.setStyle(
        TableStyle([

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.3,
                colors.grey
            ),

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor(
                    "#DCE6F1"
                )
            ),

            (
                "BACKGROUND",
                (0, -1),
                (-1, -1),
                colors.HexColor(
                    "#E8EEF5"
                )
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "LEFTPADDING",
                (0, 0),
                (-1, -1),
                1.5
            ),

            (
                "RIGHTPADDING",
                (0, 0),
                (-1, -1),
                1.5
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                2.5
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                2.5
            )

        ])
    )

    story.append(
        attendance_table
    )

    # =====================================================
    # BUILD PDF
    # =====================================================

    doc.build(
        story
    )

    output.seek(0)

    return output


# =========================================================
# EXPORT ATTENDANCE PDF
# =========================================================

@app.route(
    "/export-attendance-pdf/"
    "<site>/<int:month>/<int:year>"
)
def export_attendance_pdf(
    site,
    month,
    year
):

    site = site.upper()

    if site not in SITES:

        return "Invalid site."

    if not valid_financial_year(
        month,
        year
    ):

        return "Invalid attendance month."

    (
        employees,
        calendar_days,
        totals
    ) = get_attendance_report_data(
        site,
        month,
        year
    )

    if not employees:

        return "No employees found."

    pdf = create_attendance_report_pdf(

        site,
        month,
        year,
        employees,
        calendar_days,
        totals

    )

    filename = (

        f"{site}_Attendance_Report_"

        f"{calendar.month_name[month]}_"

        f"{year}.pdf"

    )

    return send_file(

        pdf,

        as_attachment=True,

        download_name=filename,

        mimetype="application/pdf"

    )


# =========================================================
init_db()


if __name__ == "__main__":

    app.run(

        host="0.0.0.0",

        port=5000,

        debug=True

    )

